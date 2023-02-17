from django.db import connection
from django.core.files import File
from django.contrib.auth import get_user_model
from django.contrib.auth.models import User

from .models import Info, AttachmentFile, InfoTypeChoices, InfoRelation
from eqs.models import Eqtype

from openpyxl import load_workbook
import pytz
import os

def importInfo():
    infotype = [
        InfoTypeChoices.TECHINICAL,
        InfoTypeChoices.FAILURE_CASE,
        InfoTypeChoices.SAFETY,
        InfoTypeChoices.EVENT,
    ]
    wb_info = load_workbook('test_data/info.xlsx')
    ws_info = wb_info['info']
    infos = []
    Info.objects.all().delete()
    is_new_sys = False
    if ws_info.cell(1, 5) == '内容':
        is_new_sys = True
    for row in ws_info.iter_rows(min_row=2):
        if row[5].value < 5:
            if row[3].value:
                id = row[0].value
                managerID = row[1].value
                title = row[2].value
                if is_new_sys:
                    sammary = row[3].value
                    content = row[4].value
                    info_type = row[5].value
                    updated_at = row[10].value.replace(
                        tzinfo=pytz.timezone('Asia/Tokyo'))
                    created_at = row[11].value.replace(
                        tzinfo=pytz.timezone('Asia/Tokyo'))
                    created_by = get_user_model().objects.get_or_create(
                        username=row[7].value)
                    updated_by = get_user_model().objects.get_or_create(
                        username=row[11].value)
                else:
                    sammary = row[3].value[0:511]
                    content = row[3].value
                    info_type = infotype[row[5].value - 1]
                    updated_at = row[6].value.replace(
                        tzinfo=pytz.timezone('Asia/Tokyo'))
                    created_at = row[6].value.replace(
                        tzinfo=pytz.timezone('Asia/Tokyo'))
                    created_by = User.objects.get(username='A608')
                    updated_by = User.objects.get(username='A608')

                info = Info(id=id, managerID=managerID, title=title,
                            sammary=sammary, content=content,
                            info_type=info_type, updated_at=updated_at,
                            created_at=created_at, created_by=created_by,
                            updated_by=updated_by)

                infos.append(info)
    Info.objects.bulk_create(infos)
    importInfoEqTypes()


def importInfoEqTypes():
    wb_info_eqtype = load_workbook('test_data/infoEqType.xlsx')
    ws_info_eq_type = wb_info_eqtype['infoEqType']

    infos = []
    info_eq_type = {}
    info_id = None
    eq_types = []
    for row in ws_info_eq_type.iter_rows(min_row=2):
        if info_id == row[1].value:
            eq_type = Eqtype.objects.get_or_none(
                id=row[2].value + '_' + row[3].value)
            if eq_type:
                eq_types.append(eq_type)
        else:
            if info_id:
                info_eq_type[info_id] = eq_types

            info_id = row[1].value
            eq_types = []

    info_eq_type[info_id] = eq_types
    for key, val in info_eq_type.items():
        print(key)
        if Info.objects.filter(id=int(key)).exists():
            info = Info.objects.get_or_create(id=key)
            print(info, val)
            for eq in val:
                info[0].eqtypes.add(eq)
            infos.append(info[0])


def importInfoFiles():
    AttachmentFile.objects.all().delete()
    cursor = connection.cursor()
    cursor.execute('alter table info_attachment auto_increment = 1')

    attachmentFiles = []
    wb_load_file_list = load_workbook('test_data/infoFileList.xlsx', data_only=True)
    ws_file_lsit = wb_load_file_list['filelist']
    pk = 1
    for row in ws_file_lsit.iter_rows(min_row=2):
        if not row[3].value:
            if row[6].value == 1:
                os.rename('/home/pi/django/rise_info/uploads/migratedData/info/' + str(row[2].value),'/home/pi/django/rise_info/uploads/migratedData/info/' + str(row[5].value))
        if row[6].value:
            info_pk = row[1].value
            flnm = row[5].value
            print(row, flnm, info_pk)
            attachmentFile = importInfoFile(id=pk, flnm=flnm, info_pk=info_pk)
            if attachmentFile:
                attachmentFiles.append(attachmentFile)
                pk = pk + 1
    AttachmentFile.objects.bulk_create(attachmentFiles)


def importInfoFile(flnm: str, info_pk: int, id: int):
    print('info_pk', info_pk)
    info = Info.objects.get_or_none(pk=int(info_pk))
    if info:
        file = getMigratedData(flnm)
        attachmentFile = AttachmentFile(id=id, info=info, filename=flnm)
        attachmentFile.file.save(flnm, file)
        return attachmentFile
    else:
        return None

def getMigratedData(flnm: str):
    return File(open('/home/pi/django/rise_info/uploads/migratedData/info/' + flnm, 'rb'))
