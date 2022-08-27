from venv import create
from django.db import models, connection
from django.core.files import File
from django.contrib.auth import get_user_model

from rise_info.baseModels import CommonInfo, BaseAttachment, file_upload_path, BaseCommnets
from eqs.models import Eqtype
from offices.models import Office

from datetime import date
from openpyxl import load_workbook
import pytz


def importInfo():
    infotype = [
        InfoTypeChoices.TECHINICAL,
        InfoTypeChoices.FAILURE_CASE,
        InfoTypeChoices.SAFETY,
        InfoTypeChoices.EVENT,
    ]
    wb_info = load_workbook('test_data/info.xlsx')
    ws_info = wb_info['info']
    infos = []
    Info.objects.all().delete()
    if ws_info.cell(1, 5) == '内容':
        is_new_sys = True
    for row in ws_info.iter_rows(min_row=2):
        if row[5].value < 5:
            if row[3].value:
                id = row[0].value
                managerID = row[1].value
                title = row[2].value
                disclosure_date = row[6].value
                is_disclosed = row[8].value
                if is_new_sys:
                    sammary = row[3].value
                    content = row[4].value
                    info_type = row[5].value
                    updated_at = row[10].value.replace(
                        tzinfo=pytz.timezone('Asia/Tokyo'))
                    created_at = row[11].value.replace(
                        tzinfo=pytz.timezone('Asia/Tokyo'))
                    created_by = get_user_model().objects.get_or_create(
                        username=row[7].value)
                    updated_by = get_user_model().objects.get_or_create(
                        username=row[11].value)
                    info = Info(id=id, managerID=managerID, title=title, sammary=sammary, content=content,
                                disclosure_date=disclosure_date, info_type=info_type, is_disclosed=is_disclosed,
                                updated_at=updated_at, created_at=created_at, created_by=created_by, updated_by=updated_by)
                else:
                    sammary = row[3].value[0:511]
                    content = row[3].value
                    info_type = infotype[row[5].value - 1]
                    updated_at = disclosure_date.replace(
                        tzinfo=pytz.timezone('Asia/Tokyo'))
                    info = Info(id=id, managerID=managerID, title=title, sammary=sammary, content=content,
                                disclosure_date=disclosure_date, info_type=info_type, is_disclosed=is_disclosed,
                                updated_at=updated_at)

                infos.append(info)
    Info.objects.bulk_create(infos)
    importInfoEqTypes()


def importInfoEqTypes():
    wb_info_eqtype = load_workbook('test_data/infoEqType.xlsx')
    ws_info_eq_type = wb_info_eqtype['infoEqType']

    infos = []
    info_eq_type = {}
    info_id = None
    eq_types = []
    for row in ws_info_eq_type.iter_rows(min_row=2):
        if info_id == row[1].value:
            eq_type = Eqtype.objects.get_or_none(
                id=row[2].value + '_' + row[3].value)
            if eq_type:
                eq_types.append(eq_type)
        else:
            if info_id:
                info_eq_type[info_id] = eq_types

            info_id = row[1].value
            eq_types = []

    info_eq_type[info_id] = eq_types
    for key, val in info_eq_type.items():
        print(key)
        if Info.objects.filter(id=int(key)).exists():
            info = Info.objects.get_or_create(id=key)
            print(info, val)
            for eq in val:
                info[0].eqtypes.add(eq)
            infos.append(info[0])


def importInfoFiles():
    AttachmentFile.objects.all().delete()
    cursor = connection.cursor()
    cursor.execute('alter table info_attachment auto_increment = 1')

    attachmentFiles = []
    wb_load_file_list = load_workbook('test_data/infoFileList.xlsx')
    ws_file_lsit = wb_load_file_list['filelist']
    pk = 1
    for row in ws_file_lsit.iter_rows(min_row=2):
        if row[3].value:
            info_pk = row[1].value
            flnm = row[2].value
            print(row, flnm)
            attachmentFile = importInfoFile(id=pk, flnm=flnm, info_pk=info_pk)
            attachmentFiles.append(attachmentFile)
            pk = pk + 1
    AttachmentFile.objects.bulk_create(attachmentFiles)


def importInfoFile(flnm: str, info_pk: int, id: int):
    info = Info.objects.get_or_none(pk=info_pk)
    file = getMigratedData(flnm)
    attachmentFile = AttachmentFile(id=id, info=info, filename=flnm)
    attachmentFile.file.save(flnm, file)
    return attachmentFile


def getMigratedData(flnm: str):
    return File(open('/home/pi/django/rise_info/uploads/migratedData/info/' + flnm, 'rb'))


class InfoTypeChoices(models.TextChoices):
    TECHINICAL = 'technical', '信頼性技術情報'
    FAILURE_CASE = 'failure_case', '障害事例情報'
    SAFETY = 'safety', '安全情報'
    EVENT = 'event', 'イベント情報'
    MANUAL = 'manual', '完成図書差し替え'


class Info(CommonInfo):
    info_type = models.CharField(verbose_name='情報種別', max_length=16, choices=InfoTypeChoices.choices,
                                 default=InfoTypeChoices.TECHINICAL, null=False, blank=False)
    is_rich_text = models.BooleanField(
        verbose_name='リッチテキスト有効', default=False, help_text='内容のリッチテキスト有効/無効')
    managerID = models.CharField(
        verbose_name='管理番号', default="TMC-解析-", null=False, blank=False, max_length=32)
    sammary = models.TextField(
        verbose_name='概要', default="", null=False, blank=True, max_length=512)
    is_add_eqtypes = models.BooleanField(verbose_name='装置型式特定', default=True)
    eqtypes = models.ManyToManyField(Eqtype, verbose_name='装置型式', blank=True)
    is_add_offices = models.BooleanField(verbose_name='官署特定', default=True)
    offices = models.ManyToManyField(Office, verbose_name='官署', blank=True)
    is_disclosed = models.BooleanField(verbose_name='公開', default=True)
    disclosure_date = models.DateField(
        verbose_name='公開日', default=date.today())

    def save(self, *args, **kwargs):
        super(Info, self).save(self, *args, **kwargs)

    def delete(self, *args, **kwargs):
        for attachment in self.attachmentfile_set.all():
            attachment.delete()
        return super().delete(*args, **kwargs)

    class Meta:
        db_table = 'infos'


class AttachmentFile(BaseAttachment):
    info = models.ForeignKey(
        Info, on_delete=models.CASCADE)
    upload_path = 'info'

    class Meta:
        db_table = 'info_attachment'


class InfoComments(BaseCommnets):
    info = models.ForeignKey(
        Info, related_name='infoComment', on_delete=models.CASCADE)
    upload_path = 'info_comments'

    class Meta:
        db_table = 'info_comments'

    def save(self, *args, **kwargs):
        if not self.pk:
            self.info.save()
        super(InfoComments, self).save(*args, **kwargs)
