from infos.models import Info
import openpyxl


def makeInfoSheet(ws):
    ws.title = 'info'
    ws.cell(1, 1).value = 'Info_ID'
    ws.cell(1, 2).value = '管理番号'
    ws.cell(1, 3).value = '件名'
    ws.cell(1, 4).value = '概要'
    ws.cell(1, 5).value = '内容'
    ws.cell(1, 6).value = 'InfoType_ID'
    ws.cell(1, 7).value = '発行年月日'
    ws.cell(1, 8).value = '登録者'
    ws.cell(1, 9).value = '登録状態'
    ws.cell(1, 10).value = '投稿日'
    ws.cell(1, 11).value = '更新日'
    ws.cell(1, 12).value = '更新者'

    row = 2
    for info in Info.objects.all():
        ws.cell(row, 1).value = info.pk
        ws.cell(row, 2).value = info.managerID
        ws.cell(row, 3).value = info.title
        ws.cell(row, 4).value = info.sammary
        ws.cell(row, 5).value = info.content
        ws.cell(row, 6).value = info.info_type
        ws.cell(row, 7).value = info.disclosure_date
        if info.created_by:
            user = info.created_by.username
            ws.cell(row, 8).value = user
        ws.cell(row, 9).value = info.select_register
        ws.cell(row, 10).value = info.created_at.replace(tzinfo=None)
        ws.cell(row, 11).value = info.updated_at.replace(tzinfo=None)
        if info.updated_by:
            user = info.updated_by.username
            ws.cell(row, 12).value = user
        row += 1
